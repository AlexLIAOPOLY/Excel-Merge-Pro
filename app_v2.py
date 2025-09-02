from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
from dotenv import load_dotenv

# 加载.env文件中的环境变量
load_dotenv()
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None
    print("[警告] pandas未安装，高级数据处理功能将不可用")
import datetime as dt
from models.database import db, TableData, TableSchema, UploadHistory
from models.excel_processor import UniversalExcelProcessor
from models.deepseek_api import DeepSeekAPIClient
from config import config

app = Flask(__name__)

# 根据环境变量选择配置
config_name = os.environ.get('FLASK_ENV', 'default')
app.config.from_object(config[config_name])

CORS(app)
db.init_app(app)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/workspace')
def workspace():
    """工作台页面"""
    return render_template('workspace.html')

@app.route('/progress', methods=['GET'])
def get_progress():
    """获取当前处理进度"""
    progress = UniversalExcelProcessor.get_progress()
    return jsonify({
        'success': True,
        'progress': progress
    })

@app.route('/clear-cache', methods=['POST'])
def clear_cache():
    """清理系统缓存"""
    try:
        UniversalExcelProcessor.clear_cache()
        return jsonify({
            'success': True,
            'message': '缓存已清理'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'清理缓存失败: {str(e)}'
        })

@app.route('/upload', methods=['POST'])
def upload_files():
    print("[系统] 收到文件上传请求")
    
    if 'files[]' not in request.files:
        print("[错误] 没有找到上传的文件")
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    files = request.files.getlist('files[]')
    results = []
    
    for file in files:
        if file.filename == '':
            continue
            
        if not allowed_file(file.filename):
            results.append({
                'filename': file.filename,
                'success': False,
                'message': '不支持的文件格式，请使用.xlsx或.xls文件'
            })
            continue
        
        original_filename = file.filename  # 保存原始文件名用于数据库
        secure_filename_value = secure_filename(file.filename)  # 安全文件名用于文件系统
        # 将用户上传的文件保存到 user_files 文件夹
        user_files_folder = 'user_files'
        os.makedirs(user_files_folder, exist_ok=True)
        file_path = os.path.join(user_files_folder, secure_filename_value)
        
        try:
            file.save(file_path)
            print(f"[系统] 文件保存成功: {file_path}")
            
            success, message, count, group_id = UniversalExcelProcessor.process_excel_file_with_grouping(file_path, original_filename)
            
            results.append({
                'filename': original_filename,
                'success': success,
                'message': message,
                'count': count,
                'group_id': group_id
            })
            
            # 用户文件保留在 user_files 文件夹中，不删除
            print(f"[系统] 用户文件保留在: {file_path}")
                
        except Exception as e:
            print(f"[错误] 处理文件 {original_filename} 时出错: {str(e)}")
            results.append({
                'filename': original_filename,
                'success': False,
                'message': f'处理失败: {str(e)}',
                'count': 0,
                'group_id': None
            })
    
    # 上传完成后清理重复分组（确保数据一致性）
    try:
        cleaned_count = UniversalExcelProcessor.cleanup_duplicate_groups()
        if cleaned_count > 0:
            print(f"[系统] 上传后清理了 {cleaned_count} 个重复分组")
    except Exception as e:
        print(f"[警告] 清理重复分组时出错: {str(e)}")
    
    return jsonify({'results': results})

@app.route('/data')
def get_data():
    print("[系统] 获取表格数据请求")
    try:
        from models.database import TableGroup
        
        # 获取所有分组的合并数据
        all_data = []
        all_schema = set()
        
        groups = TableGroup.query.order_by(TableGroup.updated_at.desc()).all()
        if groups:
            # 如果有分组，获取最新更新的分组数据
            latest_group = groups[0]  # 最近更新的分组
            data_records = TableData.query.filter_by(table_group_id=latest_group.id).order_by(TableData.created_at.asc()).all()
            all_data = [record.to_dict() for record in data_records]
            
            # 获取分组的表结构
            schemas = TableSchema.query.filter_by(table_group_id=latest_group.id, is_active=True).order_by(TableSchema.column_order).all()
            schema = [s.column_name for s in schemas]
        else:
            # 兼容旧数据：如果没有分组，使用原来的方式
            data = UniversalExcelProcessor.get_all_data()
            schema = UniversalExcelProcessor.get_current_schema()
            all_data = data
        
        # 计算统计信息
        stats = UniversalExcelProcessor.get_data_stats()
        
        print(f"[系统] 返回 {len(all_data)} 条记录，{len(schema)} 个列")
        return jsonify({
            'success': True,
            'data': all_data,
            'stats': stats,
            'schema': schema
        })
    except Exception as e:
        print(f"[错误] 获取数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/schema')
def get_schema():
    """获取当前表格结构"""
    try:
        schema = UniversalExcelProcessor.get_current_schema()
        return jsonify({
            'success': True,
            'schema': schema
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/update', methods=['POST'])
def update_record():
    try:
        data = request.get_json()
        record_id = data.get('id')
        update_data = {k: v for k, v in data.items() if k not in ['id', 'source_file', 'created_at', 'updated_at']}
        
        print(f"[系统] 更新记录 ID: {record_id}")
        
        success = UniversalExcelProcessor.update_record(record_id, update_data)
        
        if success:
            print(f"[系统] 记录更新成功")
            return jsonify({'success': True, 'message': '更新成功'})
        else:
            print(f"[错误] 记录更新失败")
            return jsonify({'success': False, 'message': '记录不存在'})
            
    except Exception as e:
        print(f"[错误] 更新记录时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/delete/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    try:
        print(f"[系统] 删除记录 ID: {record_id}")
        
        success = UniversalExcelProcessor.delete_record(record_id)
        
        if success:
            print(f"[系统] 记录删除成功")
            return jsonify({'success': True, 'message': '删除成功'})
        else:
            print(f"[错误] 记录删除失败")
            return jsonify({'success': False, 'message': '记录不存在'})
            
    except Exception as e:
        print(f"[错误] 删除记录时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/export')
def export_data():
    try:
        print("[系统] 开始导出数据")
        
        data = UniversalExcelProcessor.get_all_data()
        
        if not data:
            return jsonify({'success': False, 'message': '没有数据可导出'})
        
        # 获取当前表格结构
        schema = UniversalExcelProcessor.get_current_schema()
        
        if not schema:
            return jsonify({'success': False, 'message': '没有有效的表格结构'})
        
        # 系统字段列表（不应该出现在导出中）
        system_fields = {'id', 'source_file', 'created_at', 'updated_at'}
        
        # 过滤掉系统字段，只保留业务数据列
        business_columns = [col for col in schema if col not in system_fields and col.strip()]
        
        if not business_columns:
            return jsonify({'success': False, 'message': '没有可导出的业务数据列'})
        
        print(f"[系统] 业务数据列: {', '.join(business_columns)}")
        
        # 创建DataFrame，只包含业务数据列
        rows = []
        valid_data_count = 0
        
        for item in data:
            row = {}
            has_data = False
            
            # 只包含业务数据列，确保不包含任何系统字段
            for col in business_columns:
                value = item.get(col, '')
                # 清理数据：去除首尾空格，转换None为空字符串
                if value is not None:
                    value = str(value).strip()
                else:
                    value = ''
                row[col] = value
                
                # 检查是否有有效数据（非空且不是手动添加的空行）
                if value and value != '':
                    has_data = True
            
            # 只添加有实际数据的行，过滤掉完全空的行
            if has_data:
                rows.append(row)
                valid_data_count += 1
        
        if not rows:
            return jsonify({'success': False, 'message': '没有有效数据可导出'})
        
        # 创建DataFrame，确保列顺序与schema一致
        if HAS_PANDAS:
            df = pd.DataFrame(rows, columns=schema)
        else:
            # 简单的数据结构替代pandas
            df = {"rows": rows, "columns": schema}
        
        # 清理DataFrame：移除完全空的列
        df = df.dropna(axis=1, how='all')
        
        # 生成导出文件
        export_filename = f'数据表格_{dt.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        export_path = os.path.join('static/uploads', export_filename)
        
        # 确保导出目录存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 使用openpyxl写入Excel，优化格式
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表格"
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(export_path)
        
        print(f"[系统] 数据导出成功: {export_filename}, 共导出 {valid_data_count} 条有效记录")
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        print(f"[错误] 导出数据时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

def apply_enhanced_excel_styling(ws, num_columns, num_rows):
    """应用增强的Excel样式，使其与网页格式保持一致"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    
    # 定义专业的样式（突出的标题栏高亮效果）
    header_font = Font(bold=True, color="ffffff", name="Microsoft YaHei", size=13)  # 白色字体更醒目
    header_fill = PatternFill(start_color="4f46e5", end_color="3730a3", fill_type="solid")  # 深蓝色渐变背景
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(color="374151", name="Microsoft YaHei", size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 数值类型数据的对齐方式
    number_alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    
    # 精美的边框样式
    thin_border = Border(
        left=Side(border_style="thin", color="d1d5db"),
        right=Side(border_style="thin", color="d1d5db"),
        top=Side(border_style="thin", color="d1d5db"),
        bottom=Side(border_style="thin", color="d1d5db")
    )
    
    header_border = Border(
        left=Side(border_style="medium", color="1e293b"),
        right=Side(border_style="medium", color="1e293b"),
        top=Side(border_style="thick", color="1e293b"),
        bottom=Side(border_style="thick", color="1e293b")
    )
    
    # 应用表头样式
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # 应用数据行样式
    for row in range(2, num_rows + 2):  # +2 because we have header + data rows
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            
            # 智能判断数据类型并应用对应的对齐方式
            cell_value = str(cell.value) if cell.value else ""
            if cell_value and (cell_value.replace('.','').replace('-','').replace('+','').isdigit() or 
                             any(char in cell_value for char in ['¥', '$', '元', '%'])):
                cell.alignment = number_alignment
            else:
                cell.alignment = data_alignment
            
            cell.border = thin_border
            
            # 更精致的交替行背景色
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")
    
    # 智能调整列宽（优化算法）
    for col in range(1, num_columns + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # 计算列的最大内容长度，优化中文字符处理
        for row in range(1, num_rows + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_str = str(cell.value)
                # 更精确的中文字符宽度计算
                ascii_count = sum(1 for c in cell_str if ord(c) < 128)
                chinese_count = sum(1 for c in cell_str if ord(c) >= 128)
                adjusted_length = ascii_count + chinese_count * 1.8  # 中文字符宽度系数
                max_length = max(max_length, adjusted_length)
        
        # 设置列宽，考虑表头加粗效果和内容类型
        header_cell = ws.cell(row=1, column=col)
        header_length = len(str(header_cell.value)) * 1.2 if header_cell.value else 0  # 加粗字体稍宽
        
        final_width = max(10, min(max(max_length + 4, header_length + 2), 60))  # 改进的宽度计算
        ws.column_dimensions[column_letter].width = final_width
    
    # 设置更合适的行高（突出标题栏）
    ws.row_dimensions[1].height = 45  # 表头行高（更高更醒目）
    for row in range(2, num_rows + 2):
        ws.row_dimensions[row].height = 28  # 数据行高（稍高一些）
    
    # 冻结首行（便于浏览大量数据）
    ws.freeze_panes = "A2"
    
    # 优化打印和显示设置
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.orientation = 'landscape'  # 横向打印
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

@app.route('/export-all-groups')
def export_all_groups():
    """导出所有表格分组，每个分组一个工作表"""
    try:
        print("[系统] 开始导出所有表格分组")
        
        from models.database import TableGroup
        
        # 获取所有表格分组
        groups = TableGroup.query.order_by(TableGroup.created_at.asc()).all()
        
        if not groups:
            return jsonify({'success': False, 'message': '没有表格分组可导出'})
        
        # 创建工作簿
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        # 删除默认工作表
        wb.remove(wb.active)
        
        total_exported_records = 0
        exported_groups = 0
        
        # 系统字段列表（不应该出现在导出中）
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        
        for group in groups:
            # 获取该分组的数据
            data_records = TableData.query.filter_by(table_group_id=group.id).order_by(TableData.created_at.asc()).all()
            
            if not data_records:
                continue
                
            # 获取该分组的表格结构
            schemas = TableSchema.query.filter_by(table_group_id=group.id, is_active=True).order_by(TableSchema.column_order).all()
            schema_columns = [s.column_name for s in schemas]
            
            # 过滤掉系统字段
            business_columns = [col for col in schema_columns if col not in system_fields and col.strip()]
            
            if not business_columns:
                continue
            
            # 准备数据
            rows = []
            for record in data_records:
                row = {}
                has_data = False
                
                for col in business_columns:
                    value = getattr(record, col, '') if hasattr(record, col) else record.to_dict().get(col, '')
                    
                    # 清理数据
                    if value is not None:
                        value = str(value).strip()
                    else:
                        value = ''
                    row[col] = value
                    
                    # 检查是否有有效数据
                    if value and value != '':
                        has_data = True
                
                if has_data:
                    rows.append(row)
            
            if not rows:
                continue
            
            # 创建DataFrame
            if HAS_PANDAS:
                df = pd.DataFrame(rows, columns=business_columns)
            else:
                # 简单的数据结构替代pandas
                df = {"rows": rows, "columns": business_columns}
            
            # 创建工作表
            ws_name = group.group_name.replace('表格组_', '').replace('_', '-')[:31]  # Excel工作表名称长度限制
            ws = wb.create_sheet(title=ws_name)
            
            # 写入数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 增强的样式设置
            apply_enhanced_excel_styling(ws, len(business_columns), len(rows))
            
            total_exported_records += len(rows)
            exported_groups += 1
            print(f"[系统] 导出分组 '{group.group_name}': {len(rows)} 条记录")
        
        if exported_groups == 0:
            return jsonify({'success': False, 'message': '没有有效数据可导出'})
        
        # 生成有规律的导出文件名
        current_time = dt.datetime.now()
        date_str = current_time.strftime("%Y%m%d")
        time_str = current_time.strftime("%H%M%S")
        
        # 构建更有意义的文件名：分组数量+记录数量+时间戳
        export_filename = f'表格数据汇总_{exported_groups}组_{total_exported_records}条_{date_str}_{time_str}.xlsx'
        export_path = os.path.join('static/uploads', export_filename)
        
        # 确保导出目录存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 保存文件
        wb.save(export_path)
        
        print(f"[系统] 所有表格分组导出成功: {export_filename}, 共导出 {exported_groups} 个分组, {total_exported_records} 条记录")
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        print(f"[错误] 导出所有表格分组时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/history')
def get_upload_history():
    try:
        print("[系统] 获取上传历史")
        history = UploadHistory.query.order_by(UploadHistory.upload_time.desc()).limit(20).all()
        return jsonify({
            'success': True,
            'history': [h.to_dict() for h in history]
        })
    except Exception as e:
        print(f"[错误] 获取上传历史时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/uploaded-files')
def get_uploaded_files():
    """获取所有上传的文件列表"""
    try:
        print("[系统] 获取已上传文件列表")
        
        # 获取成功上传的文件历史
        files = UploadHistory.query.filter_by(status='success').order_by(UploadHistory.upload_time.desc()).all()
        
        # 为每个文件添加其数据统计
        files_with_stats = []
        for file_record in files:
            file_data = file_record.to_dict()
            
            # 统计该文件的数据记录数
            data_count = TableData.query.filter_by(source_file=file_record.filename).count()
            file_data['current_records'] = data_count
            
            # 检查文件是否还有关联的数据（可能被部分删除了）
            file_data['has_data'] = data_count > 0
            
            files_with_stats.append(file_data)
        
        print(f"[系统] 返回 {len(files_with_stats)} 个文件记录")
        return jsonify({
            'success': True,
            'files': files_with_stats,
            'total': len(files_with_stats)
        })
    except Exception as e:
        print(f"[错误] 获取文件列表时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/uploaded-files/<filename>/data')
def get_file_data(filename):
    """获取特定文件的数据"""
    try:
        print(f"[系统] 获取文件 {filename} 的数据")
        
        # 获取该文件的所有数据记录
        data_records = TableData.query.filter_by(source_file=filename).order_by(TableData.created_at.asc()).all()
        
        if not data_records:
            return jsonify({'success': False, 'message': '文件数据不存在或已被删除'})
        
        # 获取数据
        data = [record.to_dict() for record in data_records]
        
        # 从第一条记录获取schema（所有记录应该有相同的结构）
        if data:
            first_record = data[0]
            # 排除系统字段，获取业务字段作为schema
            system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
            schema = [key for key in first_record.keys() if key not in system_fields]
        else:
            schema = []
        
        # 统计信息
        stats = {
            'total_records': len(data),
            'source_file': filename,
            'total_columns': len(schema),
            'upload_time': data_records[0].created_at.strftime('%Y-%m-%d %H:%M:%S') if data_records else None
        }
        
        print(f"[系统] 返回文件 {filename} 的 {len(data)} 条记录")
        return jsonify({
            'success': True,
            'data': data,
            'schema': schema,
            'stats': stats
        })
    except Exception as e:
        print(f"[错误] 获取文件数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/uploaded-files/<filename>/export')
def export_file_data(filename):
    """导出特定文件的数据"""
    try:
        print(f"[系统] 开始导出文件 {filename}")
        
        # 获取该文件的所有数据记录
        data_records = TableData.query.filter_by(source_file=filename).order_by(TableData.created_at.asc()).all()
        
        if not data_records:
            return jsonify({'success': False, 'message': '文件数据不存在或已被删除'})
        
        # 准备数据
        data = [record.to_dict() for record in data_records]
        
        if not data:
            return jsonify({'success': False, 'message': '没有数据可导出'})
        
        # 系统字段列表
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        
        # 获取业务数据列
        first_record = data[0]
        business_columns = [col for col in first_record.keys() if col not in system_fields and col.strip()]
        
        if not business_columns:
            return jsonify({'success': False, 'message': '没有可导出的数据列'})
        
        # 准备导出数据
        rows = []
        for item in data:
            row = {}
            has_data = False
            
            for col in business_columns:
                value = item.get(col, '')
                if value is not None:
                    value = str(value).strip()
                else:
                    value = ''
                row[col] = value
                
                if value and value != '':
                    has_data = True
            
            if has_data:
                rows.append(row)
        
        if not rows:
            return jsonify({'success': False, 'message': '没有有效数据'})
        
        # 创建DataFrame
        if HAS_PANDAS:
            df = pd.DataFrame(rows, columns=business_columns)
        else:
            df = {"rows": rows, "columns": business_columns}
        
        # 生成导出文件名
        safe_filename = filename.replace('.xlsx', '').replace('.xls', '').replace(' ', '_')
        current_time = dt.datetime.now()
        date_str = current_time.strftime("%Y%m%d_%H%M%S")
        
        export_filename = f'{safe_filename}_{len(rows)}条_{date_str}.xlsx'
        export_path = os.path.join('static/uploads', export_filename)
        
        # 确保导出目录存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 使用openpyxl写入Excel
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = safe_filename[:31]
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 应用样式
        apply_enhanced_excel_styling(ws, len(business_columns), len(rows))
        
        # 保存文件
        wb.save(export_path)
        
        print(f"[系统] 文件导出成功: {export_filename}, 共导出 {len(rows)} 条记录")
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        print(f"[错误] 导出文件数据时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/uploaded-files/<filename>/delete', methods=['DELETE'])
def delete_file_data(filename):
    """删除特定文件的数据库记录（不删除物理文件，保留在工作台中）"""
    try:
        print(f"[系统] 开始删除文件 {filename} 的数据库记录（物理文件保留）")
        
        # 删除该文件的所有数据记录
        deleted_count = TableData.query.filter_by(source_file=filename).delete()
        
        # 删除上传历史记录
        UploadHistory.query.filter_by(filename=filename).delete()
        
        db.session.commit()
        
        print(f"[系统] 文件 {filename} 数据库记录删除成功，删除了 {deleted_count} 条数据记录（物理文件保留在工作台中）")
        return jsonify({
            'success': True,
            'message': f'文件删除成功，共删除 {deleted_count} 条记录'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 删除文件数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

# 旧的清空数据接口已移除，使用 /clear-all 代替

@app.route('/column/add', methods=['POST'])
def add_column():
    try:
        data = request.get_json()
        column_name = data.get('column_name', '').strip()
        insert_position = data.get('insert_position')  # 新增位置参数
        
        if not column_name:
            return jsonify({'success': False, 'message': '列名不能为空'})
        
        print(f"[系统] 添加新列: {column_name}, 位置: {insert_position}")
        
        success, message = UniversalExcelProcessor.add_column(column_name, insert_position)
        
        return jsonify({'success': success, 'message': message})
        
    except Exception as e:
        print(f"[错误] 添加列时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/column/delete', methods=['POST'])
def delete_column():
    try:
        data = request.get_json()
        column_name = data.get('column_name', '').strip()
        
        if not column_name:
            return jsonify({'success': False, 'message': '列名不能为空'})
        
        print(f"[系统] 删除列: {column_name}")
        
        success, message = UniversalExcelProcessor.delete_column(column_name)
        
        return jsonify({'success': success, 'message': message})
        
    except Exception as e:
        print(f"[错误] 删除列时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/column/rename', methods=['POST'])
def rename_column():
    try:
        data = request.get_json()
        old_name = data.get('old_name', '').strip()
        new_name = data.get('new_name', '').strip()
        
        if not old_name or not new_name:
            return jsonify({'success': False, 'message': '列名不能为空'})
        
        if old_name == new_name:
            return jsonify({'success': False, 'message': '新列名与原列名相同'})
        
        print(f"[系统] 重命名列: {old_name} -> {new_name}")
        
        success, message = UniversalExcelProcessor.rename_column(old_name, new_name)
        
        return jsonify({'success': success, 'message': message})
        
    except Exception as e:
        print(f"[错误] 重命名列时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/add_row', methods=['POST'])
def add_row():
    try:
        data = request.get_json()
        insert_after_id = data.get('insert_after_id') if data else None
        insert_position = data.get('insert_position', -1) if data else -1
        
        print(f"[系统] 添加新行，插入在ID {insert_after_id} 之后，位置: {insert_position}")
        
        # 使用更新后的add_row方法
        success, message, row_id = UniversalExcelProcessor.add_row(insert_after_id)
        
        if success:
            return jsonify({'success': True, 'message': message, 'id': row_id})
        else:
            return jsonify({'success': False, 'message': message})
        
    except Exception as e:
        print(f"[错误] 添加行时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

# ==================== 智能表格分组 API ====================

@app.route('/table-groups')
def get_table_groups():
    """获取所有表格分组"""
    try:
        from models.database import TableGroup
        groups = TableGroup.query.order_by(TableGroup.updated_at.desc()).all()
        
        groups_data = []
        for group in groups:
            group_dict = group.to_dict()
            
            # 获取列映射信息
            from models.database import ColumnMapping
            mappings = ColumnMapping.query.filter_by(table_group_id=group.id).all()
            group_dict['has_mappings'] = len(mappings) > 0
            group_dict['mapping_count'] = len(mappings)
            
            groups_data.append(group_dict)
        
        return jsonify({
            'success': True,
            'groups': groups_data,
            'total_groups': len(groups_data)
        })
        
    except Exception as e:
        print(f"[错误] 获取表格分组时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/table-groups/<int:group_id>/data')
def get_group_data(group_id):
    """获取指定分组的数据"""
    try:
        from models.database import TableGroup, TableData, TableSchema
        
        # 验证分组是否存在
        group = TableGroup.query.get(group_id)
        if not group:
            return jsonify({'success': False, 'message': '分组不存在'})
        
        # 获取分组的表结构
        schemas = TableSchema.query.filter_by(
            table_group_id=group_id, 
            is_active=True
        ).order_by(TableSchema.column_order).all()
        schema_columns = [s.column_name for s in schemas]
        
        # 获取分组的数据
        data_records = TableData.query.filter_by(
            table_group_id=group_id
        ).order_by(TableData.created_at.asc()).all()
        
        data = [record.to_dict() for record in data_records]
        
        # 统计信息
        stats = {
            'total_records': len(data),
            'source_files': len(set(record.source_file for record in data_records if record.source_file)),
            'total_columns': len(schema_columns),
            'last_update': group.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return jsonify({
            'success': True,
            'data': data,
            'schema': schema_columns,
            'stats': stats,
            'group_info': group.to_dict()
        })
        
    except Exception as e:
        print(f"[错误] 获取分组数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/table-groups/<int:group_id>/mappings')
def get_group_mappings(group_id):
    """获取指定分组的列映射关系"""
    try:
        from models.database_v2 import ColumnMapping
        
        mappings = ColumnMapping.query.filter_by(table_group_id=group_id).all()
        mappings_data = [mapping.to_dict() for mapping in mappings]
        
        # 按来源文件分组
        by_file = {}
        for mapping in mappings_data:
            file = mapping['source_file']
            if file not in by_file:
                by_file[file] = []
            by_file[file].append(mapping)
        
        return jsonify({
            'success': True,
            'mappings': mappings_data,
            'by_file': by_file,
            'total_mappings': len(mappings_data)
        })
        
    except Exception as e:
        print(f"[错误] 获取列映射时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/confirm-mapping', methods=['POST'])
def confirm_column_mapping():
    """确认列映射关系"""
    try:
        from models.database_v2 import ColumnMapping
        
        data = request.get_json()
        mapping_id = data.get('mapping_id')
        
        mapping = ColumnMapping.query.get(mapping_id)
        if not mapping:
            return jsonify({'success': False, 'message': '映射不存在'})
        
        mapping.is_confirmed = True
        db.session.commit()
        
        return jsonify({'success': True, 'message': '映射确认成功'})
        
    except Exception as e:
        print(f"[错误] 确认映射时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/table-groups/rename', methods=['POST'])
def rename_table_group():
    try:
        from models.database import TableGroup
        
        data = request.get_json()
        group_id = data.get('group_id')
        new_name = data.get('new_name', '').strip()
        
        if not group_id or not new_name:
            return jsonify({'success': False, 'message': '参数不完整'})
        
        # 查找表格分组
        table_group = TableGroup.query.get(group_id)
        if not table_group:
            return jsonify({'success': False, 'message': '表格分组不存在'})
        
        # 检查名称是否已存在（同一分组名称应该是唯一的）
        existing_group = TableGroup.query.filter(
            TableGroup.group_name == new_name,
            TableGroup.id != group_id
        ).first()
        
        if existing_group:
            return jsonify({'success': False, 'message': f'表格名称 "{new_name}" 已存在'})
        
        # 更新表格名称
        old_name = table_group.group_name
        table_group.group_name = new_name
        
        db.session.commit()
        
        print(f"[系统] 表格重命名成功: {old_name} -> {new_name}")
        return jsonify({'success': True, 'message': '表格重命名成功'})
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 重命名表格时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/table-groups/<int:group_id>/delete', methods=['DELETE'])
def delete_table_group(group_id):
    """删除指定的表格分组"""
    try:
        from models.database import TableGroup, ColumnMapping
        
        # 查找表格分组
        table_group = TableGroup.query.get(group_id)
        if not table_group:
            return jsonify({'success': False, 'message': '表格不存在'})
        
        print(f"[系统] 开始删除表格分组: {table_group.group_name}")
        
        # 删除关联的数据记录
        TableData.query.filter_by(table_group_id=group_id).delete()
        
        # 删除关联的表格结构
        TableSchema.query.filter_by(table_group_id=group_id).delete()
        
        # 删除关联的列映射
        ColumnMapping.query.filter_by(table_group_id=group_id).delete()
        
        # 删除表格分组本身
        db.session.delete(table_group)
        
        db.session.commit()
        
        print(f"[系统] 表格分组删除成功: {table_group.group_name}")
        return jsonify({'success': True, 'message': '表格删除成功'})
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 删除表格分组时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/table-groups/<int:group_id>/export')
def export_single_group(group_id):
    """导出指定表格分组"""
    try:
        from models.database import TableGroup
        
        # 验证分组是否存在
        group = TableGroup.query.get(group_id)
        if not group:
            return jsonify({'success': False, 'message': '表格不存在'})
        
        print(f"[系统] 开始导出表格分组: {group.group_name}")
        
        # 获取该分组的数据
        data_records = TableData.query.filter_by(table_group_id=group_id).order_by(TableData.created_at.asc()).all()
        
        if not data_records:
            return jsonify({'success': False, 'message': '表格中没有数据'})
            
        # 获取该分组的表格结构
        schemas = TableSchema.query.filter_by(table_group_id=group_id, is_active=True).order_by(TableSchema.column_order).all()
        schema_columns = [s.column_name for s in schemas]
        
        # 系统字段列表（不应该出现在导出中）
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        
        # 过滤掉系统字段
        business_columns = [col for col in schema_columns if col not in system_fields and col.strip()]
        
        if not business_columns:
            return jsonify({'success': False, 'message': '没有可导出的数据列'})
        
        # 准备数据
        rows = []
        for record in data_records:
            row = {}
            has_data = False
            
            for col in business_columns:
                value = getattr(record, col, '') if hasattr(record, col) else record.to_dict().get(col, '')
                
                # 清理数据
                if value is not None:
                    value = str(value).strip()
                else:
                    value = ''
                row[col] = value
                
                # 检查是否有有效数据
                if value and value != '':
                    has_data = True
            
            if has_data:
                rows.append(row)
        
        if not rows:
            return jsonify({'success': False, 'message': '没有有效数据'})
        
        # 创建DataFrame
        if HAS_PANDAS:
            df = pd.DataFrame(rows, columns=business_columns)
        else:
            # 简单的数据结构替代pandas
            df = {"rows": rows, "columns": business_columns}
        
        # 生成导出文件
        safe_group_name = group.group_name.replace('表格组_', '').replace('_', '-')
        current_time = dt.datetime.now()
        date_str = current_time.strftime("%Y%m%d_%H%M%S")
        
        export_filename = f'{safe_group_name}_{len(rows)}条_{date_str}.xlsx'
        export_path = os.path.join('static/uploads', export_filename)
        
        # 确保导出目录存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 使用openpyxl写入Excel
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        ws = wb.active
        ws.title = safe_group_name[:31]  # Excel工作表名称长度限制
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 应用样式
        apply_enhanced_excel_styling(ws, len(business_columns), len(rows))
        
        # 保存文件
        wb.save(export_path)
        
        print(f"[系统] 表格导出成功: {export_filename}, 共导出 {len(rows)} 条记录")
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        print(f"[错误] 导出表格分组时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/clear-all', methods=['POST'])
def clear_all_data():
    """清空所有数据和表格分组"""
    try:
        from models.database import TableGroup, ColumnMapping
        
        print("[系统] 开始清空所有数据...")
        
        # 删除所有数据记录
        TableData.query.delete()
        
        # 删除所有表格结构
        TableSchema.query.delete()
        
        # 删除所有表格分组
        TableGroup.query.delete()
        
        # 删除所有列映射
        ColumnMapping.query.delete()
        
        # 删除所有上传历史
        UploadHistory.query.delete()
        
        db.session.commit()
        
        print("[系统] 所有数据已清空")
        return jsonify({'success': True, 'message': '所有数据已清空'})
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 清空数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

# ==================== DeepSeek AI 智能命名 API ====================

@app.route('/api/ai-rename-table', methods=['POST'])
def ai_rename_table():
    """使用DeepSeek API为表格生成智能名称"""
    try:
        from models.database import TableGroup
        
        data = request.get_json()
        group_id = data.get('group_id')
        
        if not group_id:
            return jsonify({'success': False, 'message': '表格分组ID不能为空'})
        
        # 获取表格分组
        table_group = TableGroup.query.get(group_id)
        if not table_group:
            return jsonify({'success': False, 'message': '表格分组不存在'})
        
        print(f"[AI命名] 开始为表格分组 {table_group.group_name} 生成智能名称")
        
        # 获取表格的列名
        schemas = TableSchema.query.filter_by(
            table_group_id=group_id, 
            is_active=True
        ).order_by(TableSchema.column_order).all()
        
        if not schemas:
            return jsonify({'success': False, 'message': '表格没有有效的列结构'})
        
        column_names = [s.column_name for s in schemas]
        
        # 获取一些样本数据帮助AI理解表格内容
        sample_records = TableData.query.filter_by(table_group_id=group_id).limit(3).all()
        sample_data = []
        
        for record in sample_records:
            record_dict = record.to_dict()
            # 只保留业务数据，排除系统字段
            clean_dict = {k: v for k, v in record_dict.items() 
                         if k not in ['id', 'source_file', 'created_at', 'updated_at', 'table_group_id']}
            if clean_dict:
                sample_data.append(clean_dict)
        
        # 调用DeepSeek API生成名称
        api_client = DeepSeekAPIClient()
        success, new_name, message = api_client.generate_table_name(column_names, sample_data)
        
        if success and new_name:
            # 检查名称是否与现有名称重复
            original_new_name = new_name
            counter = 1
            while True:
                existing = TableGroup.query.filter(
                    TableGroup.group_name == new_name,
                    TableGroup.id != group_id
                ).first()
                
                if not existing:
                    break
                    
                counter += 1
                new_name = f"{original_new_name}_{counter}"
                
                if counter > 100:  # 防止无限循环
                    new_name = f"{original_new_name}_{int(dt.datetime.now().timestamp())}"
                    break
            
            # 更新表格名称
            old_name = table_group.group_name
            table_group.group_name = new_name
            db.session.commit()
            
            print(f"[AI命名] 表格重命名成功: {old_name} -> {new_name}")
            
            return jsonify({
                'success': True,
                'message': 'AI智能重命名成功',
                'old_name': old_name,
                'new_name': new_name,
                'ai_message': message
            })
        else:
            return jsonify({
                'success': False,
                'message': f'AI命名失败: {message}'
            })
            
    except Exception as e:
        print(f"[AI命名] 智能重命名出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'})

@app.route('/api/test-deepseek-connection', methods=['GET'])
def test_deepseek_connection():
    """测试DeepSeek API连接"""
    try:
        api_client = DeepSeekAPIClient()
        success, message = api_client.test_connection()
        
        return jsonify({
            'success': success,
            'message': message,
            'timestamp': dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'连接测试失败: {str(e)}',
            'timestamp': dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })


# ==================== 工作台 API ====================

@app.route('/api/workspace/files')
def get_workspace_files():
    """获取工作台文件列表 - 扫描物理文件夹"""
    try:
        import glob
        
        # 定义要扫描的文件夹
        file_folders = [
            'test_files',
            'test_files_v2',
            'user_files'  # 用户上传的文件夹
        ]
        
        files_data = []
        
        for folder in file_folders:
            if not os.path.exists(folder):
                continue
                
            # 扫描Excel文件
            excel_files = glob.glob(os.path.join(folder, '*.xlsx')) + glob.glob(os.path.join(folder, '*.xls'))
            
            for file_path in excel_files:
                try:
                    # 重新导入os模块以防冲突
                    import os as os_module
                    # 获取文件信息
                    file_name = os_module.path.basename(file_path)
                    file_stat = os_module.stat(file_path)
                    file_size = file_stat.st_size
                    created_time = dt.datetime.fromtimestamp(file_stat.st_ctime)
                    modified_time = dt.datetime.fromtimestamp(file_stat.st_mtime)
                    
                    # 查询数据库中该文件的记录数
                    records_count = TableData.query.filter_by(source_file=file_name).count()
                    
                    # 获取该文件的列信息
                    if records_count > 0:
                        # 从第一条记录获取列信息
                        first_record = TableData.query.filter_by(source_file=file_name).first()
                        if first_record:
                            data = first_record.get_data()
                            # 排除系统字段
                            system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
                            columns = [key for key in data.keys() if key not in system_fields]
                            columns_count = len(columns)
                        else:
                            columns_count = 0
                    else:
                        columns_count = 0
                    
                    # 工作台文件显示逻辑：
                    # 已导入：文件数据已存在于数据库中（在数据合并的文件管理中可见）
                    # 已存在：文件物理存在但没有数据库记录（仅在后台存在）
                    has_data_status = records_count > 0
                    
                    files_data.append({
                        'name': file_name,
                        'path': file_path,
                        'folder': folder,
                        'size': file_size,
                        'size_formatted': format_file_size(file_size),
                        'records': records_count,
                        'columns': columns_count,
                        'created_at': created_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'updated_at': modified_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'has_data': has_data_status
                    })
                    
                except Exception as e:
                    print(f"[警告] 处理文件 {file_path} 时出错: {str(e)}")
                    continue
        
        # 按修改时间倒序排列
        files_data.sort(key=lambda x: x['updated_at'], reverse=True)
        
        return jsonify({
            'success': True,
            'files': files_data,
            'total': len(files_data)
        })
        
    except Exception as e:
        print(f"[错误] 扫描文件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

def format_file_size(size_bytes):
    """格式化文件大小"""
    if size_bytes == 0:
        return "0B"
    size_names = ["B", "KB", "MB", "GB"]
    import math
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_names[i]}"

@app.route('/api/workspace/files/<filename>/data')
def get_workspace_file_data(filename):
    """获取工作台文件数据 - 基于文件名"""
    try:
        # 验证文件是否存在
        file_found = False
        file_path = None
        file_folders = ['test_files', 'test_files_v2', 'user_files']
        
        for folder in file_folders:
            potential_path = os.path.join(folder, filename)
            if os.path.exists(potential_path):
                file_path = potential_path
                file_found = True
                break
        
        if not file_found:
            return jsonify({'success': False, 'message': '文件不存在'})
        
        # 获取该文件在数据库中的数据
        data_records = TableData.query.filter_by(source_file=filename).order_by(TableData.created_at.asc()).all()
        
        if not data_records:
            return jsonify({
                'success': True,
                'data': [],
                'schema': [],
                'stats': {
                    'total_records': 0,
                    'total_columns': 0,
                    'file_size': format_file_size(os.path.getsize(file_path)),
                    'last_update': dt.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                },
                'file_info': {
                    'name': filename,
                    'path': file_path,
                    'has_data': False
                }
            })
        
        # 转换数据
        data = [record.to_dict() for record in data_records]
        
        # 获取Schema（从第一条记录）
        first_record_data = data_records[0].get_data()
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        schema_columns = [key for key in first_record_data.keys() if key not in system_fields]
        
        # 统计信息
        stats = {
            'total_records': len(data),
            'total_columns': len(schema_columns),
            'file_size': format_file_size(os.path.getsize(file_path)),
            'last_update': dt.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        }
        
        return jsonify({
            'success': True,
            'data': data,
            'schema': schema_columns,
            'stats': stats,
            'file_info': {
                'name': filename,
                'path': file_path,
                'has_data': True
            }
        })
        
    except Exception as e:
        print(f"[错误] 获取工作台文件数据时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/workspace/files/<filename>/group')
def get_workspace_file_group(filename):
    """获取工作台文件对应的表格分组"""
    try:
        from models.database import TableGroup, TableData
        
        # 查找包含此文件数据的表格分组
        # 先通过文件名查找TableData记录
        table_data = TableData.query.filter_by(source_file=filename).first()
        
        if table_data and table_data.table_group_id:
            # 获取对应的表格分组
            group = TableGroup.query.get(table_data.table_group_id)
            if group:
                # 统计该分组的记录数
                total_records = TableData.query.filter_by(table_group_id=group.id).count()
                
                return jsonify({
                    'success': True,
                    'group': {
                        'id': group.id,
                        'group_name': group.group_name,
                        'column_count': group.column_count,
                        'total_records': total_records,
                        'confidence_score': group.confidence_score,
                        'created_at': group.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                        'updated_at': group.updated_at.strftime('%Y-%m-%d %H:%M:%S')
                    }
                })
        
        return jsonify({'success': False, 'message': '未找到该文件对应的表格分组'})
        
    except Exception as e:
        print(f"[错误] 获取工作台文件分组时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/workspace/files/upload', methods=['POST'])
def upload_workspace_file():
    """上传文件到工作台"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '没有选择文件'})
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': '不支持的文件格式，请使用.xlsx或.xls文件'})
        
        # 确保user_files目录存在
        upload_folder = 'user_files'
        os.makedirs(upload_folder, exist_ok=True)
        
        # 直接使用原始文件名，不进行secure_filename处理（支持中文）
        filename = file.filename
        print(f"[调试] 使用原始文件名: {filename}")
        
        file_path = os.path.join(upload_folder, filename)
        
        # 检查文件是否已存在
        if os.path.exists(file_path):
            return jsonify({'success': False, 'message': f'文件 "{filename}" 已存在'})
        
        # 保存文件
        file.save(file_path)
        
        print(f"[系统] 文件上传成功: {filename}")
        return jsonify({
            'success': True, 
            'message': '文件上传成功',
            'filename': filename,
            'path': file_path
        })
        
    except Exception as e:
        print(f"[错误] 上传文件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/workspace/files/<filename>/rename', methods=['PUT'])
def rename_workspace_file(filename):
    """重命名物理文件"""
    try:
        data = request.get_json()
        new_name = data.get('new_name', '').strip()
        
        if not new_name:
            return jsonify({'success': False, 'message': '新文件名不能为空'})
        
        if not new_name.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '文件名必须以.xlsx或.xls结尾'})
        
        # 查找原文件
        old_file_path = None
        file_folders = ['test_files', 'test_files_v2', 'user_files']
        
        for folder in file_folders:
            potential_path = os.path.join(folder, filename)
            if os.path.exists(potential_path):
                old_file_path = potential_path
                break
        
        if not old_file_path:
            return jsonify({'success': False, 'message': '原文件不存在'})
        
        # 构建新文件路径
        folder = os.path.dirname(old_file_path)
        new_file_path = os.path.join(folder, new_name)
        
        # 检查新文件名是否已存在
        if os.path.exists(new_file_path):
            return jsonify({'success': False, 'message': f'文件名 "{new_name}" 已存在'})
        
        # 重命名物理文件
        os.rename(old_file_path, new_file_path)
        
        # 更新数据库中的source_file字段
        affected_records = TableData.query.filter_by(source_file=filename).all()
        for record in affected_records:
            record.source_file = new_name
        
        db.session.commit()
        
        print(f"[系统] 文件重命名成功: {filename} -> {new_name}")
        return jsonify({
            'success': True, 
            'message': '文件重命名成功',
            'old_name': filename,
            'new_name': new_name,
            'affected_records': len(affected_records)
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 重命名文件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/workspace/files/<filename>/delete', methods=['DELETE'])
def delete_workspace_file(filename):
    """删除物理文件和相关数据"""
    try:
        # 查找文件
        file_path = None
        file_folders = ['test_files', 'test_files_v2', 'user_files']
        
        for folder in file_folders:
            potential_path = os.path.join(folder, filename)
            if os.path.exists(potential_path):
                file_path = potential_path
                break
        
        if not file_path:
            return jsonify({'success': False, 'message': '文件不存在'})
        
        # 删除数据库中的相关记录
        deleted_records = TableData.query.filter_by(source_file=filename).delete()
        
        # 删除物理文件（只在工作台删除物理文件）
        # 检查文件是否在用户文件夹中，只删除用户上传的文件，保留测试文件
        if file_path.startswith('user_files/'):
            os.remove(file_path)
            print(f"[系统] 删除用户文件: {file_path}")
        else:
            print(f"[系统] 保留测试文件: {file_path}，只删除数据库记录")
        
        db.session.commit()
        
        print(f"[系统] 文件删除成功: {filename}，删除了 {deleted_records} 条数据记录")
        return jsonify({
            'success': True, 
            'message': f'文件删除成功，共删除 {deleted_records} 条记录',
            'deleted_records': deleted_records
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 删除文件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/workspace/files/import', methods=['POST'])
def import_workspace_file():
    """从工作台导入文件到数据库"""
    try:
        data = request.get_json()
        file_name = data.get('file_name')
        file_path = data.get('file_path')
        
        if not file_name or not file_path:
            return jsonify({'success': False, 'message': '文件名或路径不能为空'})
        
        # 检查文件是否存在
        import os as os_module
        if not os_module.path.exists(file_path):
            return jsonify({'success': False, 'message': '文件不存在'})
        
        # 检查是否已经导入过
        existing_records = TableData.query.filter_by(source_file=file_name).count()
        if existing_records > 0:
            return jsonify({'success': False, 'message': f'文件已导入，包含 {existing_records} 条记录'})
        
        # 使用Excel处理器导入文件
        success, message, count, group_id = UniversalExcelProcessor.process_excel_file_with_grouping(file_path, file_name)
        
        if success:
            # 记录上传历史
            upload_record = UploadHistory(
                filename=file_name,
                rows_imported=count,
                status='success'
            )
            
            db.session.add(upload_record)
            db.session.commit()
            
            print(f"[系统] 从工作台导入文件成功: {file_name}, 共 {count} 条记录")
            
            return jsonify({
                'success': True,
                'message': '导入成功',
                'count': count,
                'table_name': message or '未知表格'
            })
        else:
            return jsonify({'success': False, 'message': message})
            
    except Exception as e:
        db.session.rollback()
        print(f"[错误] 导入工作台文件时出错: {str(e)}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/global-search', methods=['POST'])
def global_search():
    """全局搜索所有表格数据并返回整合结果"""
    try:
        data = request.get_json()
        search_term = data.get('search_term', '').strip()
        
        if not search_term:
            return jsonify({'success': False, 'message': '搜索关键词不能为空'})
        
        print(f"[系统] 开始全局搜索: '{search_term}'")
        
        from models.database import TableGroup
        
        # 获取所有表格分组
        groups = TableGroup.query.order_by(TableGroup.updated_at.desc()).all()
        
        if not groups:
            return jsonify({
                'success': True,
                'data': [],
                'schema': [],
                'search_term': search_term,
                'total_matches': 0,
                'matched_groups': 0,
                'message': '没有可搜索的表格数据'
            })
        
        # 存储搜索结果
        all_matched_data = []
        all_columns = set()
        search_term_lower = search_term.lower()
        matched_groups_count = 0
        total_matches = 0
        
        # 系统字段列表（不参与搜索和显示）
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        
        for group in groups:
            try:
                # 获取该分组的数据
                data_records = TableData.query.filter_by(table_group_id=group.id).order_by(TableData.created_at.asc()).all()
                
                if not data_records:
                    continue
                
                # 获取该分组的表格结构
                schemas = TableSchema.query.filter_by(table_group_id=group.id, is_active=True).order_by(TableSchema.column_order).all()
                group_columns = [s.column_name for s in schemas if s.column_name not in system_fields]
                
                # 搜索匹配的记录
                group_matches = []
                for record in data_records:
                    record_data = record.to_dict()
                    is_match = False
                    
                    # 在所有业务字段中搜索
                    for col in group_columns:
                        value = record_data.get(col, '')
                        if value and str(value).lower().find(search_term_lower) != -1:
                            is_match = True
                            break
                    
                    if is_match:
                        # 添加来源信息到匹配记录
                        matched_record = {}
                        for col in group_columns:
                            matched_record[col] = record_data.get(col, '')
                        matched_record['_source_table'] = group.group_name  # 添加来源表格信息
                        matched_record['_source_file'] = record_data.get('source_file', '')  # 添加来源文件信息
                        
                        group_matches.append(matched_record)
                        all_columns.update(group_columns)
                
                if group_matches:
                    all_matched_data.extend(group_matches)
                    matched_groups_count += 1
                    total_matches += len(group_matches)
                    print(f"[系统] 在表格组 '{group.group_name}' 中找到 {len(group_matches)} 条匹配记录")
                    
            except Exception as e:
                print(f"[警告] 搜索表格组 {group.group_name} 时出错: {str(e)}")
                continue
        
        # 构建统一的表格结构
        # 优先显示最常见的列，并在末尾添加来源信息
        common_columns = list(all_columns)
        final_schema = sorted(common_columns) + ['_source_table', '_source_file']
        
        # 补全所有记录的字段（确保所有记录都有相同的字段结构）
        for record in all_matched_data:
            for col in final_schema:
                if col not in record:
                    record[col] = ''
        
        print(f"[系统] 全局搜索完成: 在 {matched_groups_count} 个表格中找到 {total_matches} 条匹配记录")
        
        return jsonify({
            'success': True,
            'data': all_matched_data,
            'schema': final_schema,
            'search_term': search_term,
            'total_matches': total_matches,
            'matched_groups': matched_groups_count,
            'stats': {
                'total_records': total_matches,
                'source_tables': matched_groups_count,
                'total_columns': len(final_schema) - 2,  # 排除来源信息列
                'is_global_search': True
            }
        })
        
    except Exception as e:
        print(f"[错误] 全局搜索时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/export-global-search', methods=['POST'])
def export_global_search_results():
    """导出全局搜索结果"""
    try:
        data = request.get_json()
        search_term = data.get('search_term', '')
        search_data = data.get('data', [])
        search_schema = data.get('schema', [])
        
        if not search_data:
            return jsonify({'success': False, 'message': '没有搜索结果可导出'})
        
        print(f"[系统] 开始导出全局搜索结果: '{search_term}'，共 {len(search_data)} 条记录")
        
        # 系统字段列表（不应该出现在导出中）
        system_fields = {'id', 'source_file', 'created_at', 'updated_at', 'table_group_id'}
        
        # 过滤掉系统字段，但保留来源信息字段
        business_columns = [col for col in search_schema if col not in system_fields and col.strip()]
        
        if not business_columns:
            return jsonify({'success': False, 'message': '没有可导出的数据列'})
        
        # 准备导出数据
        rows = []
        for item in search_data:
            row = {}
            has_data = False
            
            for col in business_columns:
                value = item.get(col, '')
                if value is not None:
                    value = str(value).strip()
                else:
                    value = ''
                row[col] = value
                
                # 检查是否有有效数据（除了来源信息字段）
                if value and value != '' and not col.startswith('_source_'):
                    has_data = True
            
            if has_data or any(col.startswith('_source_') for col in business_columns):
                rows.append(row)
        
        if not rows:
            return jsonify({'success': False, 'message': '没有有效数据可导出'})
        
        # 创建DataFrame
        if HAS_PANDAS:
            df = pd.DataFrame(rows, columns=business_columns)
        else:
            df = {"rows": rows, "columns": business_columns}
        
        # 生成导出文件名
        safe_search_term = search_term.replace(' ', '_').replace('/', '_').replace('\\', '_')
        current_time = dt.datetime.now()
        date_str = current_time.strftime("%Y%m%d_%H%M%S")
        
        export_filename = f'全局搜索结果_{safe_search_term}_{len(rows)}条_{date_str}.xlsx'
        export_path = os.path.join('static/uploads', export_filename)
        
        # 确保导出目录存在
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        # 使用openpyxl写入Excel
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"搜索结果_{safe_search_term}"[:31]  # Excel工作表名称长度限制
        
        # 写入数据
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 应用样式
        apply_enhanced_excel_styling(ws, len(business_columns), len(rows))
        
        # 特殊处理：为来源信息列添加特殊颜色
        for col_idx, col_name in enumerate(business_columns, 1):
            if col_name.startswith('_source_'):
                # 为来源信息列设置浅绿色背景
                source_fill = PatternFill(start_color="e6f3ff", end_color="e6f3ff", fill_type="solid")
                # 设置表头
                ws.cell(row=1, column=col_idx).fill = source_fill
                # 设置所有数据行
                for row_idx in range(2, len(rows) + 2):
                    ws.cell(row=row_idx, column=col_idx).fill = source_fill
        
        # 保存文件
        wb.save(export_path)
        
        print(f"[系统] 全局搜索结果导出成功: {export_filename}, 共导出 {len(rows)} 条记录")
        
        return send_file(export_path, as_attachment=True, download_name=export_filename)
        
    except Exception as e:
        print(f"[错误] 导出全局搜索结果时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})


def create_app():
    """应用程序工厂函数"""
    with app.app_context():
        # 创建数据库表
        db.create_all()
        print("[系统] 数据库初始化完成")
        
        # 确保上传目录存在
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
            print("[系统] 上传目录创建完成")
    
    return app

if __name__ == '__main__':
    app = create_app()
    # 本地开发环境
    port = int(os.environ.get('PORT', 5002))
    print(f"[系统] 启动通用表格合并系统... 端口: {port}")
    app.run(debug=os.environ.get('FLASK_ENV') != 'production', host='0.0.0.0', port=port)
else:
    # 生产环境 (Render/Gunicorn)
    app = create_app()